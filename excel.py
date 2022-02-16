import openpyxl

wb = openpyxl.load_workbook("new.xlsx",data_only=True)
ws = wb["4月"]
cell = ws["D4"]
print(cell.value)

i = 3
while(i<=10):
    subject = ws["A" + str(i)]
    if(subject.value is None):
        break
    print(subject.value)
    i = i + 1

i = 4
counter = 0

while(i <= 10):
    subject = ws["C" + str(i)]
    if(subject.value == "A"):
        counter = counter + 1
    i = i + 1
print(counter)

customer_list = []

i = 4

while(i < 10):
    ws_1 = ws["B" + str(i)]
    if(ws_1.value not in customer_list):
        customer_list.append(ws_1.value)
    i = i + 1
    print("check")
print(customer_list)

i = 4
j = 0
list_length = len(customer_list)

sum_list = []
sum = 0

print(ws["F" + str(5)].value)

while(j != list_length):
    for k in range(10):
        ws_2 = ws["F" + str(i)]
        if(ws["B" + str(i)].value == customer_list[j]):
            sum = sum + int(ws_2.value)
        i = i + 1
    sum_list.append(sum)
    j = j + 1
    i = 4

print(sum_list)

