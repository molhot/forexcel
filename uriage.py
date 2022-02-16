from sqlite3 import Row
import openpyxl

wb = openpyxl.load_workbook("売上データ.xlsx",data_only=True)

ws = wb["Sheet1"]

#顧客IDを選定する
rownumber = 4
costomerlist = []

while(ws["B" + str(rownumber)].value != None):
    subjectcell = ws["B" + str(rownumber)].value
    if(subjectcell not in costomerlist):
        costomerlist.append(subjectcell)
    rownumber = rownumber + 1
print(costomerlist)

#各企業ごとに別シートをまとめていく

costomerlist_length = len(costomerlist)
rownumber = 4
number = 0

while(number != costomerlist_length):
    wb = openpyxl.load_workbook("売上データ.xlsx",data_only=True)
    ws = wb["Sheet1"]
    listname = "list" + str(number)
    while(ws["B" + str(rownumber)].value != None):
        if(costomerlist[number] != ws["B" + str(rownumber)].value):
            ws.delete_rows(rownumber)
            rownumber = rownumber - 1
        rownumber = rownumber + 1
    rownumber = 4
    number = number + 1
    wb.save(listname + '.xlsx')

#各企業に対して請求書を作成していく

number = 0
row_origin = 4
row_copy = 12
list_8 = []
sum_8 = 0
list_10 = []
sum_10 = 0

while(number != costomerlist_length):
    listname = "list" + str(number)
    wb_1 = openpyxl.load_workbook(listname + ".xlsx",data_only=True)
    wb_2 = openpyxl.load_workbook("請求書.xlsx",data_only=True)
    ws_1 = wb_1["Sheet1"]
    ws_2 = wb_2["Sheet1"]
    while(ws_1["A" + str(row_origin)].value != None):
        ws_2["A" + str(row_copy)].value = ws_1["A" + str(row_origin)].value
        ws_2["C" + str(row_copy)].value = ws_1["D" + str(row_origin)].value
        ws_2["E" + str(row_copy)].value = ws_1["H" + str(row_origin)].value
        ws_2["F" + str(row_copy)].value = ws_1["E" + str(row_origin)].value
        ws_2["H" + str(row_copy)].value = ws_1["F" + str(row_origin)].value
        ws_2["I" + str(row_copy)].value = ws_1["G" + str(row_origin)].value
        if(ws_2["E" + str(row_copy)].value == "*"):
            list_10.append(ws_2["I" + str(row_copy)].value)
        else:
            list_8.append(ws_2["I" + str(row_copy)].value)
        row_origin = row_origin + 1
        row_copy = row_copy + 1
    list_8_length = len(list_8)
    list_10_length = len(list_10)
    while(list_8_length > 0):
        sum_8 = sum_8 + list_8[list_8_length - 1]
        list_8_length = list_8_length - 1
    while(list_10_length > 0):
        sum_10 = sum_10 + list_10[list_10_length - 1]
        list_10_length = list_10_length - 1
    ws_2["I" + str(23)].value = sum_8
    ws_2["I" + str(24)].value = sum_10
    ws_2["I" + str(25)].value = int(sum_8 / 108 * 8)
    ws_2["I" + str(26)].value = int(sum_10 / 110 * 10)
    ws_2["I" + str(27)].value =  ws_2["I" + str(23)].value + ws_2["I" + str(24)].value 
    ws_2["C" + str(9)].value =  ws_2["I" + str(27)].value
    ws_2["A" + str(3)].value = ws_1["C" + str(4)].value
    wb_2.save(str(number) + "請求書.xlsx")
    row_origin = 4
    row_copy = 12
    number = number + 1